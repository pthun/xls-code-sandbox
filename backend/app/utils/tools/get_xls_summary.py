"""Tool that summarises XLS/XLSX spreadsheets.""" 

from __future__ import annotations

import math
from datetime import date, datetime
from typing import Any, Iterable, Mapping, Optional, Tuple

from pydantic import BaseModel, Field
from openai.types.responses import FunctionToolParam

from .base import ResponseTool, ToolExecutionResult
from .filesystem import (
    InvalidToolFilePathError,
    ToolFileNotFoundError,
    ToolFileRecord,
    ToolNotFoundError,
    resolve_tool_file,
)
from .registry import registry

try:  # pragma: no cover - optional dependency
    import openpyxl  # type: ignore[import]
except ModuleNotFoundError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

try:  # pragma: no cover - optional dependency
    import xlrd  # type: ignore[import]
except ModuleNotFoundError:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]


MAX_SAMPLE_ROWS = 5


class SpreadsheetFileArgs(BaseModel):
    """Arguments accepted by the get_xls_summary tool."""

    path: str = Field(
        ...,
        description="Path (relative to the tool uploads directory) to the spreadsheet.",
    )


class WorksheetSampleRow(BaseModel):
    row_number: int
    values: list[str]


class WorksheetSummary(BaseModel):
    name: str
    header: list[str]
    header_row_number: int | None
    total_rows: int
    data_rows: int
    columns: int
    sample_rows: list[WorksheetSampleRow]


class SpreadsheetSummary(BaseModel):
    tool_id: int
    path: str
    sheet_count: int
    sheets: list[WorksheetSummary]


GET_XLS_SUMMARY_NAME = "get_xls_summary"

GET_XLS_SUMMARY_DEFINITION = FunctionToolParam(
    type="function",
    name=GET_XLS_SUMMARY_NAME,
    description="Summarise an XLS/XLSX file by reporting sheet counts and table shapes.",
    parameters=SpreadsheetFileArgs.model_json_schema(),
    strict=False,
)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _trim_row(values: Iterable[Any]) -> list[str]:
    material = list(values)
    while material and _is_blank(material[-1]):
        material.pop()
    return [_stringify(item) for item in material]


def _summarise_rows(rows: Iterable[Tuple[int, Iterable[Any]]]) -> tuple[list[str], int | None, int, int, int, list[WorksheetSampleRow]]:
    header: list[str] = []
    header_row_number: int | None = None
    data_rows = 0
    total_rows = 0
    max_columns = 0
    samples: list[WorksheetSampleRow] = []

    for row_number, raw_values in rows:
        trimmed = _trim_row(raw_values)
        if not trimmed:
            continue
        total_rows += 1
        max_columns = max(max_columns, len(trimmed))
        if header_row_number is None:
            header = trimmed
            header_row_number = row_number
            continue
        data_rows += 1
        if len(samples) < MAX_SAMPLE_ROWS:
            samples.append(WorksheetSampleRow(row_number=row_number, values=trimmed))

    calculated_total_rows = data_rows + (1 if header_row_number is not None else 0)
    # total_rows tracks non-empty rows encountered; prefer calculated value for clarity.
    total_rows = calculated_total_rows
    return header, header_row_number, total_rows, data_rows, max_columns, samples


def _summarise_xlsx(record: ToolFileRecord) -> SpreadsheetSummary:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed, cannot process XLSX files")

    path = record.path
    if not path.exists():
        msg = f"File not found on disk: {path}"
        raise ToolFileNotFoundError(msg)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[WorksheetSummary] = []
        for worksheet in workbook.worksheets:
            rows = (
                (idx, tuple(cell for cell in row))
                for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
            )
            header, header_row_number, total_rows, data_rows, max_columns, samples = _summarise_rows(rows)
            sheets.append(
                WorksheetSummary(
                    name=worksheet.title,
                    header=header,
                    header_row_number=header_row_number,
                    total_rows=total_rows,
                    data_rows=data_rows,
                    columns=max_columns,
                    sample_rows=samples,
                )
            )
    finally:
        workbook.close()

    return SpreadsheetSummary(
        tool_id=record.tool_id,
        path=str(path),
        sheet_count=len(sheets),
        sheets=sheets,
    )


def _convert_xls_cell(cell: "xlrd.sheet.Cell", datemode: int) -> Any:
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = datetime(*xlrd.xldate_as_tuple(cell.value, datemode))
        except (TypeError, ValueError):  # pragma: no cover - defensive
            return cell.value
        if dt.time() == datetime.min.time():
            return dt.date()
        return dt
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return f"#ERR({cell.value})"
    return cell.value


def _summarise_xls(record: ToolFileRecord) -> SpreadsheetSummary:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed, cannot process XLS files")

    path = record.path
    if not path.exists():
        msg = f"File not found on disk: {path}"
        raise ToolFileNotFoundError(msg)

    book = xlrd.open_workbook(path, on_demand=True)
    try:
        sheets: list[WorksheetSummary] = []
        for sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            rows = (
                (
                    row_index + 1,
                    (_convert_xls_cell(sheet.cell(row_index, col_index), book.datemode) for col_index in range(sheet.ncols)),
                )
                for row_index in range(sheet.nrows)
            )
            header, header_row_number, total_rows, data_rows, max_columns, samples = _summarise_rows(rows)
            sheets.append(
                WorksheetSummary(
                    name=sheet_name,
                    header=header,
                    header_row_number=header_row_number,
                    total_rows=total_rows,
                    data_rows=data_rows,
                    columns=max_columns,
                    sample_rows=samples,
                )
            )
            sheet.unload()
    finally:
        book.release_resources()

    return SpreadsheetSummary(
        tool_id=record.tool_id,
        path=str(path),
        sheet_count=len(sheets),
        sheets=sheets,
    )


async def _execute_get_xls_summary(
    *, tool_id: int, arguments: Optional[Mapping[str, Any]] = None
) -> ToolExecutionResult:
    args = SpreadsheetFileArgs.model_validate(arguments or {})

    try:
        record = resolve_tool_file(
            tool_id,
            path=args.path,
        )
    except (ToolNotFoundError, ToolFileNotFoundError, InvalidToolFilePathError) as exc:
        return ToolExecutionResult(success=False, output="{}", error=str(exc))

    suffix = record.path.suffix.lower()
    try:
        if suffix == ".xlsx":
            payload = _summarise_xlsx(record)
        elif suffix == ".xls":
            payload = _summarise_xls(record)
        else:
            raise ValueError(f"Unsupported spreadsheet format '{suffix}'.")
    except Exception as exc:
        return ToolExecutionResult(success=False, output="{}", error=str(exc))

    return ToolExecutionResult(
        success=True,
        output=payload.model_dump_json(),
    )


get_xls_summary_tool = ResponseTool(
    name=GET_XLS_SUMMARY_NAME,
    definition=GET_XLS_SUMMARY_DEFINITION,
    executor=_execute_get_xls_summary,
)

registry.register(get_xls_summary_tool)

__all__ = [
    "GET_XLS_SUMMARY_NAME",
    "GET_XLS_SUMMARY_DEFINITION",
    "SpreadsheetFileArgs",
    "SpreadsheetSummary",
    "get_xls_summary_tool",
]
