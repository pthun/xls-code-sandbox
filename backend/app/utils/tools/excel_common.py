"""Shared models and helpers for Excel-focused tools."""

from __future__ import annotations

import math
import os
import tempfile
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Generator, Generic, Literal, TypeVar, Union

from pydantic.generics import GenericModel

try:  # pragma: no cover - optional dependency
    import openpyxl  # type: ignore[import]
    from openpyxl.utils import (  # type: ignore[attr-defined]
        column_index_from_string,
        get_column_letter,
        quote_sheetname,
        range_boundaries,
        coordinate_to_tuple,
    )
except ModuleNotFoundError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    column_index_from_string = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]
    quote_sheetname = None  # type: ignore[assignment]
    range_boundaries = None  # type: ignore[assignment]
    coordinate_to_tuple = None  # type: ignore[assignment]


PathStr = str
SheetName = str
A1Range = str
ColumnRef = Union[int, str]
CellValue = Union[str, int, float, bool, date, datetime, Decimal, None]
CSVText = str
TableName = str
ErrorCode = Literal["NOT_FOUND", "INVALID_ARG", "IO", "CONFLICT", "INTERNAL"]

ResultStatus = Literal["ok", "error"]

T = TypeVar("T")



from .filesystem import (
    InvalidToolFilePathError,
    ToolFileNotFoundError,
    ToolFileRecord,
    ToolNotFoundError,
    resolve_tool_file,
)


class Result(GenericModel, Generic[T]):
    """Standardised envelope for all Excel tools."""

    status: ResultStatus
    data: T | None
    error_code: ErrorCode | None
    message: str | None

    @classmethod
    def ok(cls, data: T) -> "Result[T]":
        return cls(status="ok", data=data, error_code=None, message=None)

    @classmethod
    def error(cls, *, error_code: ErrorCode, message: str) -> "Result[Any]":
        return cls(status="error", data=None, error_code=error_code, message=message)




def resolve_workbook_record(
    tool_id: int,
    path: PathStr,
    *,
    folder_prefix: str | None = None,
) -> ToolFileRecord:
    """Resolve and validate a workbook record for the given tool."""

    try:
        record = resolve_tool_file(tool_id, path=path, folder_prefix=folder_prefix)
    except ValueError as exc:  # pragma: no cover - defensive
        raise InvalidToolFilePathError(str(exc)) from exc

    if not record.path.exists():
        msg = f"File not found on disk: {record.path}"
        raise ToolFileNotFoundError(msg)
    ensure_path_is_xlsx(record.path)
    return record

class WorkbookReadError(RuntimeError):
    """Raised when a workbook cannot be loaded for reading."""


class WorkbookWriteError(RuntimeError):
    """Raised when a workbook cannot be saved."""


class CoordinateError(ValueError):
    """Raised when row/column references are invalid."""


class SheetNotFoundError(LookupError):
    """Raised when a requested worksheet is missing."""


class WorkbookItemNotFoundError(LookupError):
    """Raised when a named workbook structure cannot be located."""


class ConflictError(RuntimeError):
    """Raised when an operation would conflict with existing workbook structures."""


def ensure_openpyxl_available() -> None:
    """Ensure openpyxl and helpers are importable before proceeding."""

    if openpyxl is None:
        raise WorkbookReadError("openpyxl is not installed, cannot process XLSX files")
    if column_index_from_string is None or get_column_letter is None or range_boundaries is None:
        raise WorkbookReadError("openpyxl utilities are unavailable")


def ensure_path_is_xlsx(path: Path) -> None:
    """Validate that the given path looks like an XLSX workbook."""

    if path.suffix.lower() != ".xlsx":
        msg = f"Unsupported spreadsheet format '{path.suffix}'. Only .xlsx is supported."
        raise WorkbookReadError(msg)



def get_required_sheet(workbook: "openpyxl.Workbook", sheet: SheetName):  # type: ignore[name-defined]
    """Return the requested sheet or raise a descriptive error."""

    try:
        return workbook[sheet]
    except KeyError as exc:
        raise SheetNotFoundError(f"Sheet '{sheet}' not found") from exc


@contextmanager
def open_workbook(path: Path, *, data_only: bool, read_only: bool = True) -> Generator["openpyxl.Workbook", None, None]:  # type: ignore[name-defined]
    """Open an XLSX workbook with standard validation."""

    ensure_openpyxl_available()
    ensure_path_is_xlsx(path)

    try:
        workbook = openpyxl.load_workbook(  # type: ignore[attr-defined]
            path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        msg = f"File not found on disk: {path}"
        raise WorkbookReadError(msg) from exc
    except Exception as exc:  # pragma: no cover - defensive
        msg = f"Failed to open workbook: {exc}"
        raise WorkbookReadError(msg) from exc

    try:
        yield workbook
    finally:
        workbook.close()


@contextmanager
def open_workbook_for_write(path: Path) -> Generator["openpyxl.Workbook", None, None]:  # type: ignore[name-defined]
    """Open an XLSX workbook for modification (not read-only)."""

    ensure_openpyxl_available()
    ensure_path_is_xlsx(path)

    try:
        workbook = openpyxl.load_workbook(path)  # type: ignore[attr-defined]
    except FileNotFoundError as exc:
        msg = f"File not found on disk: {path}"
        raise WorkbookWriteError(msg) from exc
    except Exception as exc:  # pragma: no cover - defensive
        msg = f"Failed to open workbook: {exc}"
        raise WorkbookWriteError(msg) from exc

    try:
        yield workbook
    finally:
        workbook.close()


def save_workbook_atomic(workbook: "openpyxl.Workbook", path: Path) -> None:  # type: ignore[name-defined]
    """Persist the workbook to disk using a temporary file for atomicity."""

    temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
    os.close(temp_fd)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    except Exception as exc:  # pragma: no cover - defensive
        raise WorkbookWriteError(f"Failed to save workbook: {exc}") from exc
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


def parse_column_ref(column: ColumnRef) -> int:
    """Return a 1-based column index from an int or Excel-style letter."""

    if isinstance(column, int):
        if column < 1:
            raise CoordinateError("Column indices must be 1-based")
        return column
    if not isinstance(column, str):
        raise CoordinateError("Column reference must be int or str")
    ensure_openpyxl_available()
    assert column_index_from_string is not None  # for type checkers
    try:
        return column_index_from_string(column.upper())
    except Exception as exc:
        raise CoordinateError(f"Invalid column reference '{column}'") from exc


def column_letter(index: int) -> str:
    """Convert a 1-based column index to its Excel letter."""

    ensure_openpyxl_available()
    if index < 1:
        raise CoordinateError("Column indices must be >= 1")
    assert get_column_letter is not None
    return get_column_letter(index)


def to_a1(row: int, column: int) -> str:
    """Return an A1-style cell coordinate."""

    if row < 1:
        raise CoordinateError("Row indices must be >= 1")
    return f"{column_letter(column)}{row}"


def parse_a1_range(a1: A1Range) -> tuple[int, int, int, int]:
    """Convert an A1 range string to integer boundaries."""

    ensure_openpyxl_available()
    assert range_boundaries is not None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(a1)
    except Exception as exc:
        raise CoordinateError(f"Invalid A1 range '{a1}'") from exc
    return min_row, min_col, max_row, max_col


def parse_cell_reference(cell_ref: str) -> tuple[int, int]:
    """Convert a single cell address (e.g. 'B3') to row and column indices."""

    ensure_openpyxl_available()
    assert coordinate_to_tuple is not None
    try:
        row, column = coordinate_to_tuple(cell_ref)
    except Exception as exc:
        raise CoordinateError(f"Invalid cell reference '{cell_ref}'") from exc
    return row, column


def normalize_row(row: int) -> int:
    """Validate and normalise a row index."""

    if row < 1:
        raise CoordinateError("Row indices must be >= 1")
    return row


def value_to_display(value: Any) -> str:
    """Convert a cell value into a safe display string."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value)


def row_values_trimmed(values: list[Any]) -> list[str]:
    """Trim trailing blank-like values and stringify for sampling."""

    trimmed = list(values)
    while trimmed and is_blank_value(trimmed[-1]):
        trimmed.pop()
    return [value_to_display(item) for item in trimmed]


def is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def sheet_dimension_bounds(sheet: "openpyxl.worksheet.worksheet.Worksheet") -> tuple[int, int, int, int]:  # type: ignore[name-defined]
    """Return numeric bounds for a worksheet's used range."""

    ensure_openpyxl_available()
    dimension = sheet.calculate_dimension(force=True)
    min_row, min_col, max_row, max_col = parse_a1_range(dimension)
    return min_row, min_col, max_row, max_col


def format_range(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    """Format numeric bounds as an A1 range."""

    start = to_a1(min_row, min_col)
    end = to_a1(max_row, max_col)
    if start == end:
        return start
    return f"{start}:{end}"


def quote_sheet(sheet: SheetName) -> str:
    """Quote a sheet name for formulas if needed."""

    ensure_openpyxl_available()
    assert quote_sheetname is not None
    return quote_sheetname(sheet)


class CsvBuilder:
    """Utility that builds CSV text with minimal dependencies."""

    def __init__(self) -> None:
        self._rows: list[list[str]] = []

    def append(self, row: list[Any]) -> None:
        self._rows.append([value_to_display(item) for item in row])

    def render(self) -> str:
        if not self._rows:
            return ""
        # Simple CSV writer with Excel defaults (comma separator, double-quote quoting)
        escaped_rows: list[str] = []
        for row in self._rows:
            escaped_cells: list[str] = []
            for cell in row:
                if any(ch in cell for ch in [",", "\n", '"']):
                    escaped = '"' + cell.replace('"', '""') + '"'
                else:
                    escaped = cell
                escaped_cells.append(escaped)
            escaped_rows.append(",".join(escaped_cells))
        return "\n".join(escaped_rows)


class FormulaCounter:
    """Helper to count formula cells within a worksheet."""

    def __init__(self) -> None:
        self.count = 0

    def feed(self, value: Any) -> None:
        if isinstance(value, str) and value.startswith("="):
            self.count += 1

    def tally(self) -> int:
        return self.count


def serialize_result(model: Result[Any]) -> str:
    """Serialize a Result payload to JSON for tool responses."""

    return model.model_dump_json()


def result_from_exception(exc: Exception) -> Result[Any]:
    """Map known exceptions to standardised error Result payloads."""

    message = str(exc) or exc.__class__.__name__
    if isinstance(exc, (ToolFileNotFoundError, ToolNotFoundError, SheetNotFoundError, WorkbookItemNotFoundError)):
        return Result.error(error_code="NOT_FOUND", message=message)
    if isinstance(exc, InvalidToolFilePathError):
        return Result.error(error_code="INVALID_ARG", message=message)
    if isinstance(exc, ConflictError):
        return Result.error(error_code="CONFLICT", message=message)
    if isinstance(exc, (WorkbookReadError, WorkbookWriteError)):
        return Result.error(error_code="IO", message=message)
    if isinstance(exc, CoordinateError):
        return Result.error(error_code="INVALID_ARG", message=message)
    return Result.error(error_code="INTERNAL", message=message)
