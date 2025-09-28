"""Tool that generates XLSX variations by mutating an existing sheet or adding new ones."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated, Any, Iterable, Literal, Mapping, Optional
from uuid import uuid4

from pydantic import BaseModel, ConfigDict, Field, ValidationError, model_validator
from openai.types.responses import FunctionToolParam

from .base import ResponseTool, ToolExecutionResult
from .filesystem import (
    ToolFileNotFoundError,
    ToolNotFoundError,
    UPLOAD_ROOT,
    db_connection,
    resolve_tool_file,
)
from .get_xls_summary import SpreadsheetFileArgs
from .registry import registry

try:  # pragma: no cover - optional dependency
    import openpyxl  # type: ignore[import]
    from openpyxl.utils import get_column_letter, range_boundaries
except ModuleNotFoundError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    range_boundaries = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]


def _stringify(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _cell_has_data(cell: "openpyxl.cell.Cell") -> bool:  # type: ignore[name-defined]
    value = cell.value
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


@dataclass(slots=True)
class TableContext:
    sheet: "openpyxl.worksheet.worksheet.Worksheet"  # type: ignore[name-defined]
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    columns: list[str]


class NewColumnSpec(BaseModel):
    """Specification for a column to add to the table."""

    model_config = ConfigDict(extra="forbid")

    name: str = Field(..., min_length=1, description="Name of the column to add.")
    values: list[Any] | None = Field(
        default=None,
        description="Optional explicit values for each data row (must match row count).",
    )
    default_value: Any | None = Field(
        default=None,
        description="Fallback value for existing rows when explicit values are not supplied.",
    )

    @model_validator(mode="after")
    def _ensure_values_or_default(self) -> "NewColumnSpec":
        if self.values is None and self.default_value is None:
            msg = "Provide either 'values' or 'default_value' for each new column"
            raise ValueError(msg)
        return self


class AddColumnsOperation(BaseModel):
    """Operation that adds one or more columns to the table."""

    model_config = ConfigDict(extra="forbid")

    operation: Literal["add_columns"] = "add_columns"  # type: ignore[assignment]
    columns: list[NewColumnSpec] = Field(
        ..., min_length=1, description="Columns to append to the table in order."
    )
    overwrite_existing: bool = Field(
        default=False,
        description="Whether to overwrite existing columns with the same name instead of failing.",
    )


class AppendRowsOperation(BaseModel):
    """Operation that appends fully specified rows to the table."""

    model_config = ConfigDict(extra="forbid")

    operation: Literal["append_rows"] = "append_rows"  # type: ignore[assignment]
    rows: list[Mapping[str, Any]] = Field(
        default_factory=list,
        description="Rows to append. Keys must match existing or newly added column names.",
    )
    fill_missing_with: Any | None = Field(
        default="",
        description="Fallback value for columns omitted in a row payload.",
    )


class CreateSheetOperation(BaseModel):
    """Operation that creates a brand-new worksheet with optional data."""

    model_config = ConfigDict(extra="forbid")

    operation: Literal["create_sheet"] = "create_sheet"  # type: ignore[assignment]
    sheet_name: str = Field(..., min_length=1)
    columns: list[str] | None = Field(
        default=None,
        description="Columns to use for the new sheet. Defaults to the base table columns when omitted.",
    )
    rows: list[Mapping[str, Any]] = Field(
        default_factory=list,
        description="Row payloads for the new sheet.",
    )
    fill_missing_with: Any | None = Field(
        default="",
        description="Fallback for columns missing in a row payload.",
    )
    based_on_table: bool = Field(
        default=True,
        description="When true and columns are omitted, reuse the base table columns for familiarity.",
    )


XlsxVariationOperation = Annotated[
    AddColumnsOperation | AppendRowsOperation | CreateSheetOperation,
    Field(discriminator="operation"),
]


class XlsxVariationArgs(SpreadsheetFileArgs):
    """Arguments accepted by the XLSX variation tool."""

    sheet_name: str = Field(..., description="Worksheet that contains the table to modify.")
    table_range: str = Field(
        ...,
        description="Excel range (e.g. 'A1:D20') describing the header and data rows to operate on.",
    )
    operations: list[XlsxVariationOperation] = Field(
        ..., min_length=1, description="Sequence of operations to apply to the workbook."
    )
    output_filename: str | None = Field(
        default=None,
        description="Optional filename for the generated workbook. Defaults to '<original>-variant.xlsx'.",
    )


class GeneratedWorkbookFile(BaseModel):
    id: int
    original_filename: str
    stored_filename: str
    path: str
    relative_path: str
    size_bytes: int
    uploaded_at: str


class XlsxVariationResult(BaseModel):
    tool_id: int
    base_file_id: int
    base_filename: str
    new_file: GeneratedWorkbookFile
    added_columns: list[str]
    appended_rows: int
    created_sheets: list[str]


GENERATE_XLSX_VARIATION_NAME = "generate_xlsx_variation"
GENERATE_XLSX_VARIATION_DEFINITION = FunctionToolParam(
    type="function",
    name=GENERATE_XLSX_VARIATION_NAME,
    description=(
        "Create a new XLSX by modifying a specified table range (adding columns/rows) and optionally "
        "adding new sheets with related data."
    ),
    parameters=XlsxVariationArgs.model_json_schema(),
    strict=False,
)


def _ensure_openpyxl_available() -> None:
    if openpyxl is None or range_boundaries is None:
        raise RuntimeError("openpyxl is not installed, cannot process XLSX files")


def _load_table_context(workbook: "openpyxl.Workbook", args: XlsxVariationArgs) -> TableContext:  # type: ignore[name-defined]
    if args.sheet_name not in workbook.sheetnames:
        msg = f"Worksheet '{args.sheet_name}' not found in workbook"
        raise ValueError(msg)

    sheet = workbook[args.sheet_name]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(args.table_range)  # type: ignore[misc]
    except ValueError as exc:
        raise ValueError(f"Invalid table_range '{args.table_range}': {exc}") from exc

    header: list[str] = []
    for col_idx in range(min_col, max_col + 1):
        value = sheet.cell(row=min_row, column=col_idx).value
        header.append(str(value).strip() if value is not None else "")

    return TableContext(
        sheet=sheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
        columns=header,
    )


def _extract_rows(context: TableContext) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(context.min_row + 1, context.max_row + 1):
        payload: dict[str, Any] = {}
        for offset, column_name in enumerate(context.columns):
            col_idx = context.min_col + offset
            cell = context.sheet.cell(row=row_idx, column=col_idx)
            payload[column_name] = cell.value if cell.value is not None else ""
        rows.append(payload)
    return rows


def _assert_columns_available(context: TableContext, additional_columns: int) -> None:
    if additional_columns <= 0:
        return
    start_col = context.max_col + 1
    end_col = context.max_col + additional_columns
    for col_idx in range(start_col, end_col + 1):
        for row_idx in range(context.min_row, context.max_row + 1):
            if _cell_has_data(context.sheet.cell(row=row_idx, column=col_idx)):
                col_letter = get_column_letter(col_idx) if get_column_letter else str(col_idx)
                msg = (
                    f"Cannot add columns: cell {col_letter}{row_idx} already contains data. "
                    "Please clear the area first."
                )
                raise ValueError(msg)


def _assert_rows_available(context: TableContext, additional_rows: int, column_count: int) -> None:
    if additional_rows <= 0:
        return
    start_row = context.max_row + 1
    end_row = context.max_row + additional_rows
    for row_idx in range(start_row, end_row + 1):
        for offset in range(column_count):
            col_idx = context.min_col + offset
            if _cell_has_data(context.sheet.cell(row=row_idx, column=col_idx)):
                col_letter = get_column_letter(col_idx) if get_column_letter else str(col_idx)
                msg = (
                    f"Cannot append rows: cell {col_letter}{row_idx} already contains data. "
                    "Please clear the area first."
                )
                raise ValueError(msg)


def _apply_add_columns(
    context: TableContext,
    data_rows: list[dict[str, Any]],
    defaults: dict[str, Any],
    op: AddColumnsOperation,
    added_columns: list[str],
) -> None:
    row_count = len(data_rows)
    for column in op.columns:
        name = column.name.strip()
        if not name:
            raise ValueError("Column names cannot be empty")
        exists = name in context.columns
        if exists and not op.overwrite_existing:
            raise ValueError(
                f"Column '{name}' already exists; set overwrite_existing=true to replace it"
            )
        values: Optional[list[Any]] = None
        if column.values is not None:
            values = [ _stringify(value) for value in column.values ]
            if len(values) != row_count:
                raise ValueError(
                    f"Column '{name}' expects {row_count} values but received {len(values)}"
                )
        default_value = _stringify(column.default_value)
        if not exists:
            context.columns.append(name)
            defaults[name] = default_value
            added_columns.append(name)
            context.max_col += 1
        else:
            defaults[name] = default_value
        for index, row in enumerate(data_rows):
            if values is not None:
                row[name] = values[index]
            else:
                row[name] = default_value


def _apply_append_rows(
    context: TableContext,
    data_rows: list[dict[str, Any]],
    defaults: dict[str, Any],
    op: AppendRowsOperation,
) -> int:
    appended = 0
    fill_value = _stringify(op.fill_missing_with)
    for incoming in op.rows:
        if not isinstance(incoming, Mapping):
            raise ValueError("Each appended row must be an object mapping column names to values")
        unknown_columns = [key for key in incoming.keys() if key not in context.columns]
        if unknown_columns:
            missing = ", ".join(unknown_columns)
            raise ValueError(f"Unknown columns for appended row: {missing}")
        row_payload: dict[str, Any] = {}
        for column in context.columns:
            if column in incoming:
                row_payload[column] = _stringify(incoming[column])
            else:
                row_payload[column] = defaults.get(column, fill_value)
        data_rows.append(row_payload)
        appended += 1
    context.max_row += appended
    return appended


def _write_table(context: TableContext, data_rows: list[dict[str, Any]]) -> None:
    # Write header
    for offset, column_name in enumerate(context.columns):
        col_idx = context.min_col + offset
        context.sheet.cell(row=context.min_row, column=col_idx, value=column_name)

    # Clear any residual values beyond the new column count within the header row range
    for col_idx in range(context.min_col + len(context.columns), context.max_col + 1):
        context.sheet.cell(row=context.min_row, column=col_idx, value=None)

    target_row = context.min_row + 1
    for row_data in data_rows:
        for offset, column_name in enumerate(context.columns):
            col_idx = context.min_col + offset
            context.sheet.cell(row=target_row, column=col_idx, value=row_data.get(column_name, ""))
        # Clear trailing cells beyond the current column count if necessary
        for col_idx in range(context.min_col + len(context.columns), context.max_col + 1):
            context.sheet.cell(row=target_row, column=col_idx, value=None)
        target_row += 1

    # Clear any residual rows up to the previous max_row to avoid stale data
    for row_idx in range(target_row, context.max_row + 1):
        for col_idx in range(context.min_col, context.max_col + 1):
            context.sheet.cell(row=row_idx, column=col_idx, value=None)

    # Update context max_row to the actual last populated row
    context.max_row = target_row - 1
    context.max_col = context.min_col + len(context.columns) - 1


def _apply_create_sheet(
    workbook: "openpyxl.Workbook",  # type: ignore[name-defined]
    base_columns: list[str],
    op: CreateSheetOperation,
    created_sheets: list[str],
) -> None:
    sheet_name = op.sheet_name.strip()
    if not sheet_name:
        raise ValueError("Sheet name cannot be empty")
    if sheet_name in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' already exists")

    columns = op.columns
    if columns is None:
        if not op.based_on_table:
            raise ValueError("Provide columns when based_on_table is false")
        columns = list(base_columns)
    if not columns:
        raise ValueError("At least one column is required for the new sheet")

    fill_value = _stringify(op.fill_missing_with)
    sheet = workbook.create_sheet(title=sheet_name)
    # Header row
    for idx, column_name in enumerate(columns, start=1):
        sheet.cell(row=1, column=idx, value=column_name)

    row_index = 2
    for row in op.rows:
        if not isinstance(row, Mapping):
            raise ValueError("Each row for the new sheet must be an object mapping column names to values")
        for idx, column_name in enumerate(columns, start=1):
            value = row[column_name] if column_name in row else fill_value
            sheet.cell(row=row_index, column=idx, value=_stringify(value))
        row_index += 1

    created_sheets.append(sheet_name)


async def _execute_generate_xlsx_variation(
    *,
    tool_id: int,
    arguments: Optional[Mapping[str, Any]] = None,
) -> ToolExecutionResult:
    try:
        args = XlsxVariationArgs.model_validate(arguments or {})
    except ValidationError as exc:
        return ToolExecutionResult(success=False, output="{}", error=str(exc))

    try:
        record = resolve_tool_file(
            tool_id,
            file_id=args.file_id,
            path=args.path,
        )
    except (ToolNotFoundError, ToolFileNotFoundError) as exc:
        return ToolExecutionResult(success=False, output="{}", error=str(exc))
    except Exception as exc:  # pragma: no cover - defensive
        return ToolExecutionResult(success=False, output="{}", error=str(exc))

    if record.path.suffix.lower() not in {".xlsx", ".xlsm"}:
        msg = f"Unsupported file type '{record.path.suffix}'. Only XLSX/XLSM files are supported."
        return ToolExecutionResult(success=False, output="{}", error=msg)

    try:
        _ensure_openpyxl_available()
    except RuntimeError as exc:
        return ToolExecutionResult(success=False, output="{}", error=str(exc))

    workbook = openpyxl.load_workbook(record.path)  # type: ignore[attr-defined]
    try:
        context = _load_table_context(workbook, args)
        data_rows = _extract_rows(context)
        defaults = {column: "" for column in context.columns}
        added_columns: list[str] = []
        appended_rows = 0
        created_sheets: list[str] = []

        # Determine required capacity expansions before mutation
        pending_new_columns = 0
        pending_new_rows = 0
        future_columns = set(name for name in context.columns if name)
        for op in args.operations:
            if isinstance(op, AddColumnsOperation):
                for column in op.columns:
                    name = column.name.strip()
                    if not name:
                        continue
                    if name not in future_columns:
                        pending_new_columns += 1
                        future_columns.add(name)
            elif isinstance(op, AppendRowsOperation):
                pending_new_rows += len(op.rows)

        _assert_columns_available(context, pending_new_columns)
        _assert_rows_available(context, pending_new_rows, len(context.columns) + pending_new_columns)

        for op in args.operations:
            if isinstance(op, AddColumnsOperation):
                _apply_add_columns(context, data_rows, defaults, op, added_columns)
            elif isinstance(op, AppendRowsOperation):
                appended_rows += _apply_append_rows(context, data_rows, defaults, op)
            elif isinstance(op, CreateSheetOperation):
                _apply_create_sheet(workbook, context.columns, op, created_sheets)
            else:  # pragma: no cover - exhaustive guard
                return ToolExecutionResult(success=False, output="{}", error="Unsupported operation type")

        _write_table(context, data_rows)

        target_dir = (UPLOAD_ROOT / str(tool_id)).resolve()
        target_dir.mkdir(parents=True, exist_ok=True)
        original_name = Path(record.original_filename)
        suggested_name = args.output_filename or f"{original_name.stem}-variant.xlsx"
        stored_filename = f"{uuid4().hex}{record.path.suffix.lower()}"
        target_path = target_dir / stored_filename
        workbook.save(target_path)
    except Exception as exc:
        workbook.close()
        return ToolExecutionResult(success=False, output="{}", error=str(exc))
    finally:
        workbook.close()

    size_bytes = target_path.stat().st_size
    uploaded_at = datetime.now(timezone.utc).isoformat()

    with db_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO tool_files (
                tool_id,
                original_filename,
                stored_filename,
                content_type,
                size_bytes,
                uploaded_at
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                tool_id,
                suggested_name,
                stored_filename,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                size_bytes,
                uploaded_at,
            ),
        )
        file_id = int(cursor.lastrowid)
        connection.commit()

    base_upload_dir = (UPLOAD_ROOT / str(tool_id)).resolve()
    try:
        relative_path = str(target_path.relative_to(base_upload_dir))
    except ValueError:  # pragma: no cover - defensive
        relative_path = stored_filename

    payload = XlsxVariationResult(
        tool_id=tool_id,
        base_file_id=record.id,
        base_filename=record.original_filename,
        new_file=GeneratedWorkbookFile(
            id=file_id,
            original_filename=suggested_name,
            stored_filename=stored_filename,
            path=str(target_path),
            relative_path=relative_path,
            size_bytes=size_bytes,
            uploaded_at=uploaded_at,
        ),
        added_columns=added_columns,
        appended_rows=appended_rows,
        created_sheets=created_sheets,
    )

    return ToolExecutionResult(
        success=True,
        output=payload.model_dump_json(),
    )


generate_xlsx_variation_tool = ResponseTool(
    name=GENERATE_XLSX_VARIATION_NAME,
    definition=GENERATE_XLSX_VARIATION_DEFINITION,
    executor=_execute_generate_xlsx_variation,
)

registry.register(generate_xlsx_variation_tool)

__all__ = [
    "GENERATE_XLSX_VARIATION_NAME",
    "GENERATE_XLSX_VARIATION_DEFINITION",
    "XlsxVariationArgs",
    "XlsxVariationResult",
    "generate_xlsx_variation_tool",
]
